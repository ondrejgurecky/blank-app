import streamlit as st
import requests
import math
import io
import openpyxl

# KONFIGURACE
API_KEY = st.secrets.get("MAPY_API")
HEADERS = {"Referer": "https://206cestovne.streamlit.app/"}
START_ADDR = "Šátalská 469/1, Praha 4, 14100 Praha"

# VOZIDLA uživatele
VOZIDLA = {
    "6AB3517": {"model": "Hyundai i30", "spotreba": 5.9, "phm": "BA95"},
    "8AA1204": {"model": "Škoda Fabia", "spotreba": 4.5, "phm": "BA95"},
    "6SR7185": {"model": "MG HS", "spotreba": 7.6, "phm": "BA95"}
}

# SAZBY MPSV 2016-2026 (oficiální vyhlášky)
PHM_CENY = {  # BA95 – cena dle MPSV vyhlášky (Kč/l)
    2016: 29.70, 2017: 29.50, 2018: 30.50, 2019: 33.10, 2020: 32.00,
    "2021a": 27.80, "2021b": 33.80,
    "2022a": 37.10, "2022b": 44.50,
    2023: 41.20, 2024: 38.20, 2025: 35.80, 2026: 34.70
}

SAZBY_KM_PERIOD = {  # rok nebo period → sazba km
    2016: 3.80, 2017: 3.90, 2018: 4.00, 2019: 4.10, 2020: 4.20,
    "2021a": 4.40, "2021b": 4.40,
    "2022a": 4.70, "2022b": 4.70,
    2023: 5.20, 2024: 5.60, 2025: 5.80, 2026: 5.90
}

VYHLASKY_PERIOD = {
    2016: "385/2015 Sb.", 2017: "440/2016 Sb.", 2018: "463/2017 Sb.", 2019: "333/2018 Sb.",
    2020: "358/2019 Sb.",
    "2021a": "589/2020 Sb.", "2021b": "589/2020 Sb.",
    "2022a": "511/2021 Sb.", "2022b": "511/2021 Sb.",
    2023: "467/2022 Sb.", 2024: "398/2023 Sb.", 2025: "475/2024 Sb.", 2026: "573/2025 Sb."
}

# Mapování period na zobrazovaný rok (pro větu)
PERIOD_ROK = {
    "2021a": 2021, "2021b": 2021,
    "2022a": 2022, "2022b": 2022,
}

# Volby v selectboxu
ROK_VOLBY = [
    2026, 2025, 2024, 2023,
    "2022b", "2022a",
    "2021b", "2021a",
    2020, 2019, 2018, 2017, 2016
]
ROK_LABELS = {
    "2022b": "2022  (14.5.–31.12.)",
    "2022a": "2022  (1.1.–13.5.)",
    "2021b": "2021  (19.10.–31.12.)",
    "2021a": "2021  (1.1.–18.10.)",
}


def cz(cislo, des=2):
    formatted = f"{cislo:,.{des}f}"
    formatted = formatted.replace(",", " ")
    formatted = formatted.replace(".", ",")
    return formatted


def geocode(adresa, api_key):
    url = "https://api.mapy.cz/v1/geocode"
    params = {"apikey": api_key, "query": adresa, "limit": 1}
    r = requests.get(url, params=params, headers=HEADERS, timeout=10)
    data = r.json()
    item = data["items"][0]
    return item["position"]["lon"], item["position"]["lat"]


def get_route(start_addr, end_addr, api_key):
    slon, slat = geocode(start_addr, api_key)
    elon, elat = geocode(end_addr, api_key)
    url = "https://api.mapy.cz/v1/routing/route"
    params = {
        "apikey": api_key,
        "start": f"{slon},{slat}",
        "end": f"{elon},{elat}",
        "routeType": "car_fast",
        "lang": "cs"
    }
    r = requests.get(url, params=params, headers=HEADERS, timeout=10)
    data = r.json()
    km = data["length"] / 1000
    min_ = data["duration"] / 60
    return km, min_


def vypocitej(adresa, spz, period, den=None, mes=None):
    """Spočítá náhrady pro danou adresu, SPZ a období."""
    try:
        km_jedno, min_jedno = get_route(START_ADDR, adresa, API_KEY)
    except Exception:
        return None

    tam_zpet_km = km_jedno * 2
    tam_zpet_min = min_jedno * 2

    rok = PERIOD_ROK.get(period, period)
    sazba = SAZBY_KM_PERIOD[period]
    phm_cena = PHM_CENY[period]
    spotreba = VOZIDLA[spz]["spotreba"]
    model = VOZIDLA[spz]["model"]

    zakladni = round(tam_zpet_km * sazba, 2)
    phm_litr = (tam_zpet_km / 100) * spotreba
    phm_nahrada = round(phm_litr * phm_cena, 2)
    celkem = math.ceil(zakladni + phm_nahrada)

    ctvrt_hodin = round(tam_zpet_min / 15) if rok < 2026 else None
    pul_hodin = round(tam_zpet_min / 30) if rok >= 2026 else None
    hod = int(tam_zpet_min // 60)
    min_ = int(tam_zpet_min % 60)

    return {
        "rok": rok, "adresa": adresa, "tam_zpet_km": tam_zpet_km,
        "model": model, "spotreba": spotreba, "sazba": sazba,
        "phm_cena": phm_cena, "zakladni": zakladni, "phm_litr": phm_litr,
        "phm_nahrada": phm_nahrada, "celkem": celkem,
        "ctvrt_hodin": ctvrt_hodin, "pul_hodin": pul_hodin,
        "hod": hod, "min_": min_,
        "pracovnici": st.session_state.get("pracovnici_radio", 1),
        "vyhlaska": VYHLASKY_PERIOD[period],
        "den": st.session_state.get("den_single", 1),
        "mes": st.session_state.get("mes_single", 1),
    }


def vygeneruj_pune(r):
    rok = r["rok"]
    adresa = r["adresa"]
    km = r["tam_zpet_km"]
    model = r["model"]
    spotreba = r["spotreba"]
    sazba = r["sazba"]
    phm_cena = r["phm_cena"]
    zakladni = r["zakladni"]
    phm_litr = r["phm_litr"]
    phm_nahrada = r["phm_nahrada"]
    celkem = r["celkem"]
    hod = r["hod"]
    min_ = r["min_"]
    pracovnici = r["pracovnici"]
    vyhlaska = r["vyhlaska"]
    den = r.get("den", 1)
    mes = r.get("mes", 1)
    datum = f"{den}. {mes}. {rok}"

    if pracovnici == 1:
        ucastnil = "Šetření se účastnil"
        prac_text = "jeden pracovník soudního exekutora"
    elif pracovnici == 2:
        ucastnil = "Šetření se účastnili"
        prac_text = "dva pracovníci soudního exekutora"
    else:
        ucastnil = "Šetření se účastnili"
        prac_text = "tři pracovníci soudního exekutora"
    prac_text2 = "1 pracovník" if pracovnici == 1 else ("2 pracovníci" if pracovnici == 2 else "3 pracovníci")

    if rok >= 2026:
        jednotky = r["pul_hodin"]
        sazba_casu = 150
        max_na_pracovnika = 1000
        jednotka_text = "půlhodin"
        jednotka_text1 = "půlhodinu"
    else:
        jednotky = r["ctvrt_hodin"]
        sazba_casu = 50
        max_na_pracovnika = 500
        jednotka_text = "čtvrthodin"
        jednotka_text1 = "čtvrthodinu"

    nahrada_na_pracovnika = min(jednotky * sazba_casu, max_na_pracovnika)
    nahrada_cas_celkem = nahrada_na_pracovnika * pracovnici

    if hod == 0:
        cas_text = f"{min_} minut"
    else:
        cas_text = f"{hod} hodin a {min_:02d} minut"

    veta = (
        f"Dne {datum} bylo provedeno místní šetření na adrese {adresa}. "
        f"Cesta ze sídla soudního exekutora a zpět činila {cz(km, 0)} km. "
        f"Při provedení výjezdu bylo využito osobní vozidlo {model}. "
        f"Dle technického průkazu činí kombinovaná spotřeba {cz(spotreba, 1)} l/100 km. "
        f"Pohonnou hmotou vozidla je Benzin 95. "
        f"Dle vyhlášky Ministerstva práce a sociálních věcí č. {vyhlaska} činí sazba základní náhrady "
        f"za 1 km jízdy {cz(sazba, 2)} Kč a výše průměrné ceny Benzinu 95 činí {cz(phm_cena, 2)} Kč. "
        f"S ohledem na výše uvedené má soudní exekutor nárok na základní náhradu ve výši "
        f"{cz(zakladni)} Kč ({cz(km, 0)} km × {cz(sazba, 2)} Kč) "
        f"a náhradu za spotřebované pohonné hmoty ve výši {cz(phm_nahrada)} Kč "
        f"({cz(km, 0)} km × {cz(spotreba, 1)} l/100 km × {cz(phm_cena, 2)} Kč), "
        f"celkem tedy zaokrouhleno na {cz(celkem, 0)} Kč. "
        + (f"S ohledem na zákonný limit činí výše určené náhrady cestovních výdajů za tuto cestu 1 500 Kč. "
           if celkem > 1500 else "")
        + f"Cesta trvala celkem {cas_text}, "
        f"bylo tedy započato {jednotky} {jednotka_text}. "
        f"Náhrada za ztrátu času činí {cz(sazba_casu, 0)} Kč za započatou {jednotka_text1}. "
        f"{ucastnil} {prac_text}. "
        f"Vzhledem k tomuto má soudní exekutor nárok na náhradu za ztrátu času, "
        f"která činí {cz(nahrada_cas_celkem, 0)} Kč."
    )
    return veta, nahrada_cas_celkem


# ─── STREAMLIT UI ────────────────────────────────────────────────────────────

st.set_page_config(page_title="Exekutorský kalkulátor", layout="wide")
st.title("🛣️ Exekutorský kalkulátor cestovních náhrad 2016–2026")
st.markdown("**Šátalská 469/1, Praha 4 → [adresa] a zpět**")

tab1, tab2 = st.tabs(["📍 Jedna adresa", "📊 Hromadné zpracování (Excel)"])

# ═══════════════════════════════════════════════════════════════
# TAB 1 – JEDNA ADRESA
# ═══════════════════════════════════════════════════════════════
with tab1:
    col1, col2, col3 = st.columns(3)
    adresa = col1.text_input("Cílová adresa", "")
    spz = col2.selectbox(
        "Vozidlo",
        list(VOZIDLA.keys()),
        format_func=lambda x: VOZIDLA[x]["model"],
        key="spz_single"
    )
    rok = col3.selectbox(
        "Rok / Období",
        ROK_VOLBY,
        format_func=lambda x: ROK_LABELS.get(x, str(x)),
        key="rok_single"
    )

    if st.button("🧮 SPOČÍTAT", type="primary", key="btn_single"):
        with st.spinner("Hledám optimální trasu..."):
            try:
                # DEBUG – dočasně
                st.write("API klíč načten:", bool(API_KEY))
                url = "https://api.mapy.cz/v1/geocode"
                params = {"apikey": API_KEY, "query": adresa, "limit": 1}
                r = requests.get(url, params=params, headers=HEADERS, timeout=10)
                st.write("HTTP status:", r.status_code)
                st.write("Odpověď API:", r.json())
                # konec DEBUG
                km_jedno, min_jedno = get_route(START_ADDR, adresa, API_KEY)
                st.session_state["trasa"] = {
                    "adresa": adresa,
                    "km_jedno": km_jedno,
                    "min_jedno": min_jedno,
                }
            except Exception as e:
                st.warning(f"🌐 API chyba: {str(e)[:80]}… Trasu se nepodařilo načíst.")

    # Přepočet náhrad při každém renderu (vozidlo/rok se mohlo změnit)
    if "trasa" in st.session_state:
        t = st.session_state["trasa"]
        tam_zpet_km = t["km_jedno"] * 2
        tam_zpet_min = t["min_jedno"] * 2

        rok_int = PERIOD_ROK.get(rok, rok)
        sazba = SAZBY_KM_PERIOD[rok]
        phm_cena = PHM_CENY[rok]
        spotreba = VOZIDLA[spz]["spotreba"]
        model = VOZIDLA[spz]["model"]

        zakladni = round(tam_zpet_km * sazba, 2)
        phm_litr = (tam_zpet_km / 100) * spotreba
        phm_nahrada = round(phm_litr * phm_cena, 2)
        celkem = math.ceil(zakladni + phm_nahrada)

        ctvrt_hodin = round(tam_zpet_min / 15) if rok_int < 2026 else None
        pul_hodin = round(tam_zpet_min / 30) if rok_int >= 2026 else None
        hod = int(tam_zpet_min // 60)
        min_ = int(tam_zpet_min % 60)

        r = {
            "rok": rok_int, "adresa": t["adresa"], "tam_zpet_km": tam_zpet_km,
            "model": model, "spotreba": spotreba, "sazba": sazba,
            "phm_cena": phm_cena, "zakladni": zakladni, "phm_litr": phm_litr,
            "phm_nahrada": phm_nahrada, "celkem": celkem,
            "ctvrt_hodin": ctvrt_hodin, "pul_hodin": pul_hodin,
            "hod": hod, "min_": min_,
            "pracovnici": st.session_state.get("pracovnici_radio", 1),
            "vyhlaska": VYHLASKY_PERIOD[rok],
            "den": st.session_state.get("den_single", 1),
            "mes": st.session_state.get("mes_single", 1),
        }
        st.session_state["vysledky"] = r

        st.divider()

        col1, col2 = st.columns(2)
        with col1:
            st.metric("📏 Vzdálenost tam + zpět", f"{cz(r['tam_zpet_km'], 0)} km")
            st.metric("⏱️ Doba jízdy", f"{r['hod']}:{r['min_']:02d} h")
            st.metric("💰 Náhrada km + PHM", f"{cz(r['celkem'], 0)} Kč")
        with col2:
            st.markdown("**Detail:**")
            st.write(f"*Základní:* **{cz(r['zakladni'])} Kč** ({cz(r['sazba'], 2)} Kč/km)")
            st.write(f"*PHM:* **{cz(r['phm_nahrada'])} Kč** ({cz(r['phm_litr'], 2)} l × {cz(r['phm_cena'], 2)} Kč/l)")
            if r["rok"] >= 2026:
                st.write(f"*Půlhodiny:* **{r['pul_hodin']}** × 150 Kč (max 1 000 Kč/pracovník)")
                st.metric("⏱️ Počet půlhodin", r["pul_hodin"])
            else:
                st.write(f"*Čtvrthodiny:* **{r['ctvrt_hodin']}** × 50 Kč (max 500 Kč/pracovník)")
                st.metric("⏱️ Počet čtvrthodin", r["ctvrt_hodin"])

        st.warning("⚠️ **Exekuční limit: max 1 500 Kč/cestu**")

        st.divider()
        st.subheader("📄 Věta pro PUNE")

        pracovnici = st.radio(
            "Počet pracovníků soudního exekutora:",
            options=[1, 2, 3],
            format_func=lambda x: {1: "1 pracovník", 2: "2 pracovníci", 3: "3 pracovníci"}[x],
            horizontal=True,
            key="pracovnici_radio"
        )

        col_d, col_m = st.columns([1, 3])
        den = col_d.number_input("Den šetření", min_value=1, max_value=31, value=1, step=1, key="den_single")
        mes = col_m.selectbox("Měsíc šetření", list(range(1, 13)), key="mes_single")

        r["pracovnici"] = pracovnici
        r["den"] = den
        r["mes"] = mes

        veta, nahrada_cas = vygeneruj_pune(r)

        st.info(veta)
        st.caption(f"💼 Náhrada za ztrátu času celkem: **{cz(nahrada_cas, 0)} Kč**")
        st.markdown("*Zkopírujte text níže:*")
        st.code(veta, language=None)

# ═══════════════════════════════════════════════════════════════
# TAB 2 – HROMADNÉ ZPRACOVÁNÍ EXCELU
# ═══════════════════════════════════════════════════════════════
with tab2:
    st.markdown("Nahrajte soubor `.xls`/`.xlsx` se **seznamem adres ve sloupci A** (bez záhlaví). Aplikace doplní:")
    st.markdown("- **Sloupec B** – celková náhrada km + PHM (zaokrouhleno, Kč)")
    st.markdown("- **Sloupec C** – počet čtvrthodin / půlhodin")

    col1, col2, col3 = st.columns(3)
    spz_batch = col1.selectbox(
        "Vozidlo",
        list(VOZIDLA.keys()),
        format_func=lambda x: VOZIDLA[x]["model"],
        key="spz_batch"
    )
    rok_batch = col2.selectbox(
        "Rok / Období",
        ROK_VOLBY,
        format_func=lambda x: ROK_LABELS.get(x, str(x)),
        key="rok_batch"
    )

    uploaded = st.file_uploader("Nahrajte Excel soubor", type=["xls", "xlsx"])

    if uploaded and st.button("🧮 SPOČÍTAT HROMADNĚ", type="primary", key="btn_batch"):
        wb = openpyxl.load_workbook(uploaded)
        ws = wb.active

        # Přidej záhlaví do B1 a C1 pokud jsou prázdné
        jednotka_hlavicka = "půlhodin" if rok_batch >= 2026 else "čtvrthodin"
        if ws["B1"].value is None:
            ws["B1"] = "Náhrada km+PHM (Kč)"
        if ws["C1"].value is None:
            ws["C1"] = f"Počet {jednotka_hlavicka}"

        adresy = []
        for row in ws.iter_rows(min_row=1):
            val = row[0].value
            if val:
                adresy.append((row[0].row, str(val).strip()))

        progress = st.progress(0, text="Zpracovávám adresy...")
        chyby = []

        for i, (row_idx, adresa_batch) in enumerate(adresy):
            progress.progress((i + 1) / len(adresy), text=f"({i+1}/{len(adresy)}) {adresa_batch}")
            vysl = vypocitej(adresa_batch, spz_batch, rok_batch)
            if vysl is None:
                ws.cell(row=row_idx, column=2).value = "CHYBA"
                ws.cell(row=row_idx, column=3).value = "CHYBA"
                chyby.append(adresa_batch)
            else:
                ws.cell(row=row_idx, column=2).value = vysl["celkem"]
                jednotky = vysl["pul_hodin"] if rok_batch >= 2026 else vysl["ctvrt_hodin"]
                ws.cell(row=row_idx, column=3).value = jednotky

        progress.empty()

        # Uložit do paměti a nabídnout ke stažení
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.success(f"✅ Zpracováno {len(adresy) - len(chyby)}/{len(adresy)} adres.")
        if chyby:
            st.warning(f"⚠️ Nepodařilo se zpracovat {len(chyby)} adres: {', '.join(chyby)}")

        st.download_button(
            label="⬇️ Stáhnout výsledky (.xlsx)",
            data=output,
            file_name="cestovne_vysledky.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ─── INFO PANEL ──────────────────────────────────────────────────────────────
with st.expander("ℹ️ O aplikaci"):
    st.markdown("""
    **Funkce:**
    - Mapy.cz API routy (tam-zpět)
    - MPSV sazby 2016–2026
    - Vozidla: Fabia 4.5l, i30 5.9l, MG HS 7.6l
    - Čtvrthodiny (do 2025, max 500 Kč/prac.) + půlhodiny (od 2026, max 1 000 Kč/prac.)
    - Generátor věty pro PUNE s volbou počtu pracovníků
    - Hromadné zpracování adres z Excelu

    **Deploy:** `pip install streamlit requests openpyxl`, `streamlit run app.py`
    """)

st.caption("🎯 Exekutorský úřad Mgr. Jana Škarpy, Šátalská 469/1, Praha 4")
